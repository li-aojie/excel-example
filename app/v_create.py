from openpyxl.workbook import Workbook


def create_excel(file_path: str):
    wb = Workbook()
    del wb['Sheet']  # 删除初始Sheet表
    sheet = wb.create_sheet(title='新建')
    sheet.cell(row=1, column=1, value='测试')
    wb.save(file_path)
    return


create_excel(r'C:\Users\jack\Desktop\test.xlsx')
